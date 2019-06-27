# -*- coding: utf-8 -*-
from datetime import datetime
from openpyxl import load_workbook, Workbook

wb = load_workbook('courses.xlsx')
students_sheet = wb['students']
time_sheet = wb['time']

def conbine():
    conbine_sheet = wb.create_sheet(title='conbine')
    conbine_sheet.append(['创建时间','课程名称','学习人数','学习时间'])
    for stu in students_sheet.values:
        if stu[2] != '学习人数':
            for time in time_sheet.values:
                if time[1] == stu[1]:
                    conbine_sheet.append(list(stu) + [time[2]])
    wb.save('courses.xlsx')

def split():
    conbine_sheet = wb['conbine']
    years = set()
    for item in conbine_sheet.values:
        if item[0] != '创建时间':
            years.add(item[0].strftime('%Y'))   
    for year in years:
        wb_temp = Workbook()
        ws = wb_temp.active
        ws.append(['创建时间','课程名称','学习人数','学习时间'])
        for item_by_year in conbine_sheet.values:
            if item_by_year[0] != '创建时间':
                if item_by_year[0].strftime('%Y') == year:
                    ws.append(item_by_year)
        wb_temp.save('{}.xlsx'.format(year))


if __name__ == '__main__':
    conbine()
    split()
